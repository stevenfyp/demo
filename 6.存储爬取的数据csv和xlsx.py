#csv写入
# import csv
# csv_file = open('demo.csv','w',newline='',encoding='utf-8')
# writer =csv.writer(csv_file)
# writer.writerow(['电影','豆瓣评分'])
# writer.writerow(['银河护卫队','8.0'])
# writer.writerow(['复仇者联盟','8.1'])
# csv_file.close()


#csv读取
import csv
csv_file = open('demo.csv','r',newline='')
reader=csv.reader(csv_file)
for row in reader:
    print(row)


# #excel
# import openpyxl
# wb=openpyxl.Workbook()
# sheet=wb.active
# sheet.title='new title'
# sheet['A1'] = '漫威宇宙'
# rows= [['美国队长','钢铁侠','蜘蛛侠'],['是','漫威','宇宙', '经典','人物']]
# for i in rows:
#     sheet.append(i)
# print(rows)
# wb.save('Marvel.xlsx')





写入的代码xlsx
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '漫威'
rows = ['1','2','3'],['美国队长','钢铁侠','蜘蛛侠'],['漫威','宇宙', '经典']
for x in rows:
    sheet.append(x)
wb.save('漫威.xlsx')

#读取代码
import openpyxl
wb = openpyxl.load_workbook('漫威.xlsx')
sheetname = wb.sheetnames
print(sheetname)
for x in sheetname:
    sheet = wb[x]
    A1_cell = sheet['A1']
    A1_value = A1_cell.value
    print(A1_value)
